import calendar
from openpyxl.styles import Font
from calendar import monthrange, Calendar
import pyperclip
import datetime
import time
import openpyxl
from tkinter import *
import keyboard
from project_class import Project
from data import Data
from ui import AppInterface
import os
import csv
import locale


# Current date and time.
date = datetime.datetime.now()
month = date.month
months = ["Leden", "Únor", "Březen", "Duben", "Květen", "Červen", "Červenec", "Srpen", "Září", "Říjen", "Listopad", "Prosinec"]
month_name = months[month-1]
year = date.year
first_day_of_the_month = date.replace(day=1)
locale.setlocale(locale.LC_ALL, "fr_FR.UTF-8")

#--------------------------INIT-----------------------------------
file_path = "odpisy hodin Major.xlsx"
main_workbook = openpyxl.load_workbook("odpisy hodin Major.xlsx")
# active_sheet = main_workbook.active
current_sheet_name = f"{month_name}{year}"

#--- Check for data save file ---#
filename = "project_data.txt"

if not os.path.exists(filename):
    with open(filename, mode="w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["project", "time"])



#----------------------NEW MONTH SHEET--------------------------------------

def get_workdays(year, month, first_day):
    num_days = calendar.monthrange(year, month)[1]
    days = [datetime.date(year, month, day) for day in range(1, num_days+1)]
    workdays = []
    for x in days:
        if x.weekday()==0 or x.weekday()==1 or x.weekday()==2 or x.weekday()==3 or x.weekday()==4:
            workdays.append(x)
    return workdays

# If sheet exists, skip, if new month, create new.
try:
    active_sheet = main_workbook[current_sheet_name]

except KeyError:
    print("novy_sesit")
    #New Month -> clone previous sheet
    source = main_workbook.active
    new_sheet = main_workbook.copy_worksheet(source)
    new_sheet.title = current_sheet_name

    #reset cells
    month_anchor = first_day_of_the_month.weekday()+3 #sets initial row for the first day of the month.
    active_sheet = main_workbook[new_sheet.title]
    active_sheet["A1"] = f"Odpisy za měsíc {month_name} {year}"

    default_font = "000000"
# Delete cells
    for col in range(3, 49):#49 is the end
        for row in range(2, 33):
            active_sheet.cell(row, col).value = None
            cell = active_sheet.cell(row=row, column=col)
            cell.font = Font(color=default_font)
    for col in range(3, 49):
        cell = active_sheet.cell(row=33, column=col)
        cell.font = Font(color=default_font)

    #Create signatures and others columns:
    active_sheet.cell(2, 47).value = "ostatni"
    active_sheet.cell(2, 48).value = "signature"

    day = 0
    workdays = get_workdays(year=year, month=month, first_day=first_day_of_the_month)
# Write Dates
    for row in range(3, 32):
        active_sheet.cell(row, 2).value = None
        if row >= month_anchor and row != 8 and row != 14 and row != 20 and row != 26 and day < len(workdays):
            active_sheet.cell(row, 2).value = workdays[day].strftime("%d.%m.%Y")
            day += 1
    main_workbook.save(file_path)

#------------------------------------MAIN----------------------------------------

data_class = Data()
project_class = Project()
interface = AppInterface(project_class, data_class)



#----------------------PROJECT, HOURS ENTRY--------------------------------------

active_sheet = main_workbook[current_sheet_name]



try: #This is ---- last project entry to the save file ----
    project_class.stop_time()
    data_class.write_data(project_class.project_name, project_class.project_time)

    try:
        project_data = data_class.read_data()
    except FileNotFoundError:
        pass

    # This is ---- Adding all projects and hours to the excel ----
    for project in project_data:
        print(project["project"])
        print(project["time"])

        active_project = project["project"]
        time_spent = float(project["time"])
        projects = []
        dates = []
        for col in range(3, 49):
            projects.append(active_sheet.cell(2, col).value)
        for row in range(3,31):
            dates.append(active_sheet.cell(row, 2).value)

        try:
            date_index = dates.index(date.strftime("%d.%m.%Y"))
            last_date_index = date_index
        except ValueError:
            date_index = last_date_index + 1

        if active_project in projects:
            project_index = projects.index(active_project)

            current_value = active_sheet.cell(date_index+3, project_index+3).value
            try:
                new_value = current_value + time_spent
            except TypeError:
                new_value = time_spent
            active_sheet.cell(date_index + 3, project_index + 3, value=locale.format_string("%.2f", new_value, grouping=True))
        else:
            project_index = 0
            for idx, item in enumerate(projects):
                if item == None:
                    project_index = idx
                    break
            print(project_index)
            print(projects)

            active_sheet.cell(date_index + 3, project_index + 3, value=locale.format_string("%.2f", time_spent, grouping=True))
            active_sheet.cell(2, project_index + 3).value = active_project

    # Delete save file after relation safely ends
    data_class.delete_data()

# save changes and exit
    main_workbook.save(file_path)

# When exit program, just pass out.
except AttributeError:
    print("Exited without action")
    pass

