import datetime
import openpyxl
from project_class import Project
from data import Data
from ui import AppInterface
from sheet_proc import CopyStyle
import os
import csv
import locale
import json


# Function to read configuration
def read_configuration():
    default_config = {
        "excel_location": ".",
        "default_projects": ["Others", "RnD"],
        "normalize_hours": 7.5,
        "language": 0,  # Default language (0 - English)
        "name": "projectLog"
    }

    if os.path.isfile('config.json'):
        with open('config.json', 'r') as config_file:
            return json.load(config_file)
    else:
        return default_config


# --------------------------INIT-----------------------------------


# Read the configuration
config = read_configuration()

# Set the locale based on the language configuration
if config["language"] == 1:  # 1 for Czech
    locale.setlocale(locale.LC_ALL, "cs_CZ.UTF-8")
else:  # Default to English
    locale.setlocale(locale.LC_ALL, "en_US.UTF-8")
    

# Current date and time.
date = datetime.datetime.now()
month = date.month

# Set the locale based on the language configuration, setup config variables
if config["language"] == 1:  # 1 for Czech
    locale.setlocale(locale.LC_ALL, "cs_CZ.UTF-8")
    months = [
        "Leden", "Únor", "Březen", "Duben", "Květen", "Červen", 
        "Červenec", "Srpen", "Září", "Říjen", "Listopad", "Prosinec"
    ]
if config["language"] == 0:  # Default to English
    locale.setlocale(locale.LC_ALL, "en_US.UTF-8")
    months = [
        "January", "February", "March", "April", "May", "June", 
        "July", "August", "September", "October", "November", "December"
    ]


month_name = months[month - 1]
year = date.year
first_day_of_the_month = date.replace(day=1)
locale.setlocale(locale.LC_ALL, "fr_FR.UTF-8")

config["TABLE_WIDTH"] = config.get("table_width", 46)
config["FIRST_COLUMN"] = 3
config["LAST_COLUMN"] = config["FIRST_COLUMN"] + config["TABLE_WIDTH"]
config["DATE_COLUMN"] = 2
config["PROJECT_ROW"] = 2

print(config["excel_location"])
normalized_path = os.path.normpath(config["excel_location"])
print(normalized_path)
TEMPLATE_PATH = "Project_Sheet.xlsx"
FILE_PATH = os.path.join(normalized_path, config["name"] + ".xlsx")

current_sheet_name = f"{month_name}{year}"
sheet_processor = CopyStyle(
    template = TEMPLATE_PATH, 
    file_path = FILE_PATH, 
    current_sheet_name = current_sheet_name, 
    first_day_of_the_month = first_day_of_the_month, 
    month_name = month_name, 
    year = year, 
    month = month,
    config = config
    )


# --- Check for data save file ---#
filename = "project_data.txt"
expected_line = "project,time\n"

if not os.path.exists(filename):
    with open(filename, mode="w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["project", "time"])

# Open the text file and read the first line
with open(filename, "r") as file:
    lines = file.readlines()

if not lines or lines[0] != expected_line:
    print("debug")
    with open(filename, "w") as file:
        file.write(expected_line)  # Write the correct 0th line
        file.writelines(lines)  # Write the original content



# --------------------------------NEW MONTH SHEET---------------------------------------



# If sheet exists, skip, if new month, create new.

try:
    print(FILE_PATH)
    print(f"opening {month_name} month in {FILE_PATH} sheet")
    main_workbook = openpyxl.load_workbook(FILE_PATH)
    active_sheet = main_workbook[current_sheet_name]
    main_workbook.close()


except FileNotFoundError:
    # New Month -> clone from template
    print(f"{FILE_PATH} not found. Creating new workbook")
    sheet_processor.produce_workbook()


except KeyError:
    # New Month -> clone previous sheet
    print(f"creating {month_name} month sheet")
    sheet_processor.produce_worksheet()


# -----------------------------------MAIN LOOP---------------------------------------


if __name__ == "__main__":
    data_class = Data()
    project_class = Project()
    interface = AppInterface(project_class, data_class, config)


# ----------------------PROJECT, HOURS ENTRY--------------------------------------


try:  # This is ---- last project entry to the save file ----
    main_workbook = openpyxl.load_workbook(FILE_PATH)
    active_sheet = main_workbook[current_sheet_name]
    project_class.stop_time()
    data_class.write_data(project_class.project_name, project_class.project_time)

    try:
        project_data = data_class.read_data()
    except FileNotFoundError:
        pass

    # This is ---- Adding all projects and hours to the excel ----
    for project in project_data:
        print(project["project"])
        print(project["time"])

        active_project = project["project"]
        time_spent = float(project["time"])
        projects = []
        dates = []
        for col in range(config["FIRST_COLUMN"], config["LAST_COLUMN"]):
            projects.append(active_sheet.cell(config["PROJECT_ROW"], col).value)
        for row in range(3, 38):    # Static range for one month
            dates.append(active_sheet.cell(row, 2).value)

        try:
            date_index = dates.index(date.strftime("%d.%m.%Y"))
            last_date_index = date_index
        except ValueError:
            date_index = last_date_index + 1

        if active_project in projects:
            project_index = projects.index(active_project)

            current_value = active_sheet.cell(date_index + 3, project_index + 3).value
            try:
                new_value = current_value + time_spent
            except TypeError:
                new_value = time_spent
            # active_sheet.cell(date_index + 3, project_index + 3, value=float(locale.format_string("%.1f", new_value, grouping=False)))
            active_sheet.cell(
                date_index + 3, project_index + 3, value=round(new_value, 1)
            )
        else:
            project_index = 0
            for idx, item in enumerate(projects):
                if item == None:
                    project_index = idx
                    break
            print(project_index)
            print(projects)

            # active_sheet.cell(date_index + 3, project_index + 3, value=float(locale.format_string("%.1f", time_spent, grouping=False)))
            active_sheet.cell(
                date_index + 3, project_index + 3, value=round(time_spent, 1)
            )
            active_sheet.cell(2, project_index + 3).value = active_project


    # Delete save file after relation safely ends
    data_class.delete_data()

    # save changes and exit
    main_workbook.save(FILE_PATH)
    main_workbook.close()
    

# When exit program, just pass out.
except AttributeError:
    print("Exited without action")
    pass

# If exited with normalization step, normalize hours
if interface.normalize_on_exit == True:
    print("debug")
    sheet_processor.normalize_hours(col_start=config["FIRST_COLUMN"], col_end=config["LAST_COLUMN"], target_hours=config["normalize_hours"]) 

interface.normalize_on_exit == False