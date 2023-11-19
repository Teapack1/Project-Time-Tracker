import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from copy import copy 
import calendar
import datetime

class CopyStyle:
    def __init__(self, template, file_path, current_sheet_name, first_day_of_the_month, month_name, year, month):
        # Initialize the CopyStyle object with paths and sheet name
        self.template_path = template
        self.file_path = file_path
        self.current_sheet_name = current_sheet_name
        self.first_day_of_the_month = first_day_of_the_month
        self.month_name = month_name
        self.year = year
        self.month = month
        

    def copy_cell(self, source_cell, target_cell):
        # Copy the value and style from the source cell to the target cell
        target_cell.value = source_cell.value
        if source_cell.has_style:
            # Copying font, border, fill, number format, protection, and alignment
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = source_cell.number_format
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)
            
            
    def get_workdays(self, year, month):
        num_days = calendar.monthrange(year, month)[1]
        days = [datetime.date(year, month, day) for day in range(1, num_days + 1)]
        workdays = []
        for x in days:
            if (
                x.weekday() == 0
                or x.weekday() == 1
                or x.weekday() == 2
                or x.weekday() == 3
                or x.weekday() == 4
            ):
                workdays.append(x)
        return workdays


    def produce_worksheet(self):
        # Load the template workbook and select the active sheet
        template_workbook = openpyxl.load_workbook(self.template_path)
        template_sheet = template_workbook.active

        # Create a new workbook and remove the default sheet
        self.main_workbook = openpyxl.load_workbook(self.file_path)

        # Create a new sheet with the specified name
        self.new_sheet = self.main_workbook.create_sheet(title=self.current_sheet_name)

        # Copy data and formatting from the template sheet to the new sheet
        for row in template_sheet.iter_rows():
            for cell in row:
                new_cell = self.new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                self.copy_cell(cell, new_cell)

        # Copy merged cell ranges from the template to the new sheet
        for merge_cell in template_sheet.merged_cells:
            self.new_sheet.merge_cells(str(merge_cell))

        # Copy row heights and column widths from the template to the new sheet
        for row in template_sheet.row_dimensions:
            self.new_sheet.row_dimensions[row].height = template_sheet.row_dimensions[row].height
        for col in template_sheet.column_dimensions:
            self.new_sheet.column_dimensions[col].width = template_sheet.column_dimensions[col].width

        template_workbook.close()
        
        self.reset_cells()


    def produce_workbook(self):
        # Load the template workbook and select the active sheet
        template_workbook = openpyxl.load_workbook(self.template_path)
        template_sheet = template_workbook.active

        # Create a new workbook and remove the default sheet
        self.main_workbook = openpyxl.Workbook()
        self.main_workbook.remove(self.main_workbook.active)

        # Create a new sheet with the specified name
        self.new_sheet = self.main_workbook.create_sheet(title=self.current_sheet_name)

        # Copy data and formatting from the template sheet to the new sheet
        for row in template_sheet.iter_rows():
            for cell in row:
                new_cell = self.new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                self.copy_cell(cell, new_cell)

        # Copy merged cell ranges from the template to the new sheet
        for merge_cell in template_sheet.merged_cells:
            self.new_sheet.merge_cells(str(merge_cell))

        # Copy row heights
        for row in template_sheet.row_dimensions:
            if row in template_sheet.row_dimensions:
                self.new_sheet.row_dimensions[row].height = template_sheet.row_dimensions[row].height

        # Copy column widths
        for col_letter, dim in template_sheet.column_dimensions.items():
            self.new_sheet.column_dimensions[col_letter] = copy(dim)

        template_workbook.close()
        
        self.reset_cells()
        
    
            
    def reset_cells(self):
        
        # reset cells
        first_day_weekday = self.first_day_of_the_month.weekday()
        # get what day is the first day
        month_anchor = (first_day_weekday + 4) if first_day_weekday < 5 else 4
        # sets initial row for the first day of the month.
        active_sheet = self.main_workbook[self.new_sheet.title]
        active_sheet["A1"] = f"Odpisy za měsíc {self.month_name} {self.year}"

        default_font = "000000"
        # Delete cells
        for col in range(3, 49):  # 49 is the end
            for row in range(2, 33):
                active_sheet.cell(row, col).value = None
                cell = active_sheet.cell(row=row, column=col)
                cell.font = Font(color=default_font)
        for col in range(3, 49):
            cell = active_sheet.cell(row=33, column=col)
            cell.font = Font(color=default_font)

        # Create signatures and others columns:
        active_sheet.cell(2, 47).value = "ostatni"
        active_sheet.cell(2, 48).value = "signature"

        day = 0
        workdays = self.get_workdays(year=self.year, month=self.month)
        # Write Dates
        for row in range(3, 33):
            active_sheet.cell(row, 2).value = None
            if (
                row >= month_anchor
                and row != 9
                and row != 15
                and row != 21
                and row != 27
                and day < len(workdays)
            ):
                active_sheet.cell(row, 2).value = workdays[day].strftime("%d.%m.%Y")
                day += 1
                

        self.main_workbook.save(self.file_path)
        self.main_workbook.close()
        
        
    def normalize_hours(self, row_start = 3, row_end = 31, col_start = 3, col_end = 49, target_hours=7.5):

        # Load the workbook
        workbook = openpyxl.load_workbook(self.file_path)
        sheet = workbook[self.current_sheet_name]
        
 
        for row in range(row_start, row_end + 1):
            row_values = []
            total_hours = 0
            # Calculate the sum of hours in the row
            for col in range(col_start, col_end + 1):
                cell_value = sheet.cell(row, col).value
                if isinstance(cell_value, (int, float)) and cell_value not in (None, ''):
                    total_hours += cell_value
                    row_values.append((col, cell_value))
            
            # Normalize only if the total is less than 7.5 and greater than 0
            if total_hours >= target_hours or total_hours == 0:
                continue
            
            # Calculate the factor to increase each cell value to reach the target hours
            increase_factor = target_hours / total_hours

            # Update the cell values proportionally
            for col, value in row_values:
                new_value = round(value * increase_factor, 1)
                sheet.cell(row, col).value = new_value
                
        # Save and close the workbook
        workbook.save(self.file_path)
        workbook.close()

        print("Hours have been normalized to 7.5 for each workday.")